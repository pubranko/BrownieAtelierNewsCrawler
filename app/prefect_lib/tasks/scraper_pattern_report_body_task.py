import pandas as pd
from typing import Any
from prefect import task, get_run_logger
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart.bar_chart import BarChart
from openpyxl.worksheet.worksheet import Worksheet

from prefect_lib.flows.scraper_pattern_report_const import ScraperPatternReportConst
from prefect_lib.data_models.scraper_pattern_report_data import ScraperPatternReportData


@task
def scraper_pattern_report_body_task(scraper_pattern_report_data: ScraperPatternReportData, worksheet: Worksheet):
    '''
    スクレイパー情報解析レポート用Excelの編集
    '''
    logger = get_run_logger()   # PrefectLogAdapter

    result_df: pd.DataFrame = scraper_pattern_report_data.result_df

    # 罫線定義
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)

    # エクセルシートの編集開始行数初期化
    base_row_idx: int = 3
    # 列ごとにエクセルに編集
    for col_idx, col_info in enumerate(ScraperPatternReportConst.SCRAPER_PATTERN_ANALYSIS_COLUMNS_INFO):
        for row_idx, value in enumerate(result_df[col_info[ScraperPatternReportConst.COL]]):
            # 更新対象のセル
            any: Any = worksheet[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}']
            target_cell: Cell = any

            # 更新対象のセルに値を設定
            worksheet[target_cell.coordinate] = value

            # 同値カラー調整
            if ScraperPatternReportConst.EQUIVALENT_COLOR in col_info:
                # 比較用の１つ上のセルと同じ値の場合は文字色を変更
                any: Any = worksheet[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx - 1)}']
                compare_cell: Cell = any
                if target_cell.value == compare_cell.value:
                    target_cell.font = Font(
                        color=col_info[ScraperPatternReportConst.EQUIVALENT_COLOR])

    # 各明細行のセルに罫線を設定する。
    max_cell: str = get_column_letter(
        worksheet.max_column) + str(worksheet.max_row)  # "BC55"のようなセル番地を生成
    rows = worksheet[f'a3:{max_cell}']
    if type(rows) is tuple:
        for cells in rows:
            cells: Any
            for cell in cells:
                cell.border = border

    # 列ごとに次の処理を行う。
    # 最大幅を確認
    # それに合わせた幅を設定する。
    for col in worksheet.iter_cols():
        max_length = 0
        column = col[0].column_letter  # 列名A,Bなどを取得
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        # 型ヒントでcolumn_dimensionsが存在しないものとみなされエラーが出るため、動的メソッドの実行形式で記述
        # getattr(worksheet, 'column_dimensions')()[column].width = (max_length + 2.2)
        worksheet.column_dimensions[column].width = (max_length + 5.0)

    # ウィンドウ枠の固定。２行２列は常に表示させる。
    worksheet.freeze_panes = 'c3'
