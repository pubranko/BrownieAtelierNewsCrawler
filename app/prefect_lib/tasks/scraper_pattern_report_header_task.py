from typing import Any
from prefect import task, get_run_logger
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.chart.bar_chart import BarChart
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

from prefect_lib.flows.scraper_pattern_report_const import ScraperPatternReportConst


@task
def scraper_pattern_report_header_task(worksheet: Worksheet):
    '''スクレイパー情報解析レポート用Excelの見出し編集'''
    logger = get_run_logger()   # PrefectLogAdapter

    worksheet.title = 'Scraper Info By Domain'  # シート名変更

    # 罫線定義
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    fill1 = PatternFill(patternType='solid', fgColor='0066CC')
    fill2 = PatternFill(patternType='solid', fgColor='0099CC')

    for i, col_info in enumerate(ScraperPatternReportConst.SCRAPER_PATTERN_ANALYSIS_COLUMNS_INFO):
        # 見出し１行目
        any: Any = worksheet[f'{get_column_letter(i + 1)}{str(1)}']
        head1_cell: Cell = any
        worksheet[head1_cell.coordinate] = col_info[ScraperPatternReportConst.HEAD1]
        head1_cell.fill = fill1
        head1_cell.border = border
        head1_cell.alignment = Alignment(
            horizontal="centerContinuous")  # 選択範囲内中央寄せ

        # 見出し２行目
        any: Any = worksheet[f'{get_column_letter(i + 1)}{str(2)}']
        head2_cell: Cell = any
        worksheet[head2_cell.coordinate] = col_info[ScraperPatternReportConst.HEAD2]
        head2_cell.fill = fill2
        head2_cell.border = border
        head2_cell.alignment = Alignment(horizontal="center")  # 中央寄せ
