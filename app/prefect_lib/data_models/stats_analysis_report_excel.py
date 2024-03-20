import pandas as pd
from copy import deepcopy
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Final
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.chart.bar_chart import BarChart
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from prefect_lib.data_models.stats_info_collect_data import StatsInfoCollectData


class StatsAnalysisReportExcel:
    """"""

    ############
    # 定数定義
    ############
    # 作成するレポートの見出し項目
    HEAD1: Final[str] = "head1"
    HEAD2: Final[str] = "head2"
    COL: Final[str] = "col"
    EQUIVALENT_COLOR: Final[str] = "equivalent_color"
    WARNING_VALUE: Final[str] = "warning_value"
    WARNING_BACKGROUND_COLOR: Final[str] = "warning_background_color"
    DIGIT_ADJUSTMENT: Final[str] = "digit_adjustment"
    NUMBER_FORMAT: Final[str] = "number_format"
    WARNING_VALUE_OVER: Final[str] = "warning_value_over"

    ROBOTS_RESPONSE_STATUS: Final[str] = "robots_response_status"
    DOWNLOADER_RESPONSE_STATUS: Final[str] = "downloader_response_status"

    # 各シート名
    WORKSHEET_1_NAME: Final[str] = "Stats Analysis Report"  # シート名変更
    WORKSHEET_2_NAME: Final[str] = "robots Analysis Report"  # シート名変更
    WORKSHEET_3_NAME: Final[str] = "downloader Analysis Report"  # シート名変更

    robots_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # col                       : 必須 : データフレーム列名
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {HEAD1: "集計日〜", COL: "aggregate_base_term"},
        {HEAD1: "スパイダー名", COL: "spider_name", EQUIVALENT_COLOR: "bbbbbb"},
        {
            HEAD1: "status",
            COL: "robots_response_status",
            WARNING_VALUE: [],
            WARNING_BACKGROUND_COLOR: "FF8C00",
        },
        {HEAD1: "count", COL: "count"},
    ]

    downloader_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # col                       : 必須 : データフレーム列名
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {HEAD1: "集計日〜", COL: "aggregate_base_term"},
        {HEAD1: "スパイダー名", COL: "spider_name", EQUIVALENT_COLOR: "bbbbbb"},
        {
            HEAD1: "status",
            COL: "downloader_response_status",
            WARNING_VALUE: [],
            WARNING_BACKGROUND_COLOR: "FF8C00",
        },
        {HEAD1: "count", COL: "count"},
    ]

    stats_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # head2                     : 必須 : 見出し２行目
        # col                       : 必須 : データフレーム列名
        # digit_adjustment          : 任意 : 単位の調整。1000とした場合、'value/1000'となる。
        # number_format             : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_value_over        : 任意 : 超過したらワーニングとする値
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {HEAD1: "集計日〜", HEAD2: "", COL: "aggregate_base_term"},
        {HEAD1: "スパイダー名", HEAD2: "", COL: "spider_name", EQUIVALENT_COLOR: "bbbbbb"},
        {
            HEAD1: "ログレベル件数",
            HEAD2: "CRITICAL",
            COL: "log_count/CRITICAL",
            WARNING_VALUE_OVER: 0,
            WARNING_BACKGROUND_COLOR: "FF8C00",
        },  # 0件を超えた場合ワーニング
        {
            HEAD1: "",
            HEAD2: "ERROR",
            COL: "log_count/ERROR",
            WARNING_VALUE_OVER: 0,
            WARNING_BACKGROUND_COLOR: "FF8C00",
        },  # 0件を超えた場合ワーニング
        {HEAD1: "", HEAD2: "WARNING", COL: "log_count/WARNING"},
        {
            HEAD1: "処理時間(秒)",
            HEAD2: "最小",
            COL: "elapsed_time_seconds_min",
            NUMBER_FORMAT: "#,##0.00",
        },
        {
            HEAD1: "",
            HEAD2: "最大",
            COL: "elapsed_time_seconds_max",
            NUMBER_FORMAT: "#,##0.00",
            WARNING_VALUE_OVER: 600,
            WARNING_BACKGROUND_COLOR: "FF8C00",
        },  # 処理時間が１０分を超える場合ワーニング
        {
            HEAD1: "",
            HEAD2: "合計",
            COL: "elapsed_time_seconds",
            NUMBER_FORMAT: "#,##0.00",
        },
        {
            HEAD1: "",
            HEAD2: "平均",
            COL: "elapsed_time_seconds_mean",
            NUMBER_FORMAT: "#,##0.00",
        },
        {
            HEAD1: "メモリ使用量(kb)",
            HEAD2: "最小",
            COL: "memusage/max_min",
            DIGIT_ADJUSTMENT: 1000,
        },
        {
            HEAD1: "",
            HEAD2: "最大",
            COL: "memusage/max_max",
            DIGIT_ADJUSTMENT: 1000,
            WARNING_VALUE_OVER: 500000000,
            WARNING_BACKGROUND_COLOR: "FF8C00",
        },  # メモリ使用量の500mbを超えている場合ワーニング
        {HEAD1: "", HEAD2: "平均", COL: "memusage/max_mean", DIGIT_ADJUSTMENT: 1000},
        {HEAD1: "総リクエスト数", HEAD2: "最小", COL: "downloader/request_count_min"},
        {HEAD1: "", HEAD2: "最大", COL: "downloader/request_count_max"},
        {HEAD1: "", HEAD2: "合計", COL: "downloader/request_count"},
        {
            HEAD1: "",
            HEAD2: "平均",
            COL: "downloader/request_count_mean",
            NUMBER_FORMAT: "#,##0.00",
        },
        {HEAD1: "総レスポンス数", HEAD2: "最小", COL: "downloader/response_count_min"},
        {HEAD1: "", HEAD2: "最大", COL: "downloader/response_count_max"},
        {HEAD1: "", HEAD2: "合計", COL: "downloader/response_count"},
        {
            HEAD1: "",
            HEAD2: "平均",
            COL: "downloader/response_count_mean",
            NUMBER_FORMAT: "#,##0.00",
        },
        {HEAD1: "リクエストの", HEAD2: "深さ最大", COL: "request_depth_max_max"},
        {
            HEAD1: "レスポンス量(kb)",
            HEAD2: "合計",
            COL: "downloader/response_bytes",
            DIGIT_ADJUSTMENT: 1000,
        },
        {
            HEAD1: "",
            HEAD2: "平均",
            COL: "downloader/response_bytes_mean",
            DIGIT_ADJUSTMENT: 1000,
        },
        {HEAD1: "リトライ件数", HEAD2: "合計", COL: "retry/count"},
        {HEAD1: "", HEAD2: "平均", COL: "retry/count_mean", NUMBER_FORMAT: "#,##0.00"},
        {
            HEAD1: "保存件数",
            HEAD2: "合計",
            COL: "item_scraped_count",
            WARNING_VALUE: [0],
            WARNING_BACKGROUND_COLOR: "FF8C00",
        },  # 保存件数がゼロの場合ワーニング
        {
            HEAD1: "",
            HEAD2: "平均",
            COL: "item_scraped_count_mean",
            NUMBER_FORMAT: "#,##0.00",
        },
    ]

    stats_warning_flg: bool = False
    robots_warning_flg: bool = False
    downloder_warning_flg: bool = False

    ###################
    # クラス変数定義
    ###################
    workbook: Workbook
    worksheet_1: Worksheet
    worksheet_2: Worksheet
    worksheet_3: Worksheet

    def __init__(
        self,
        stats_info_collect_data: StatsInfoCollectData,
        datetime_term_list: list[tuple[datetime, datetime]],
    ) -> None:
        # 集計結果解析レポート用のワークブックの新規作成
        self.workbook = Workbook()

        #####################################
        # スパイダー統計解析レポートを編集
        #####################################
        any: Any = self.workbook.active  # アクティブなワークシートを選択
        self.worksheet_1: Worksheet = any
        # 見出し編集
        self.stats_analysis_report_header()
        # データ編集

        # 入力パラメータの集計期間単位ごとにpandasによる解析を実行
        spider_result_all_df = stats_info_collect_data.stats_analysis_exec(
            datetime_term_list
        )

        self.stats_warning_flg: bool = self.stats_analysis_report_body(
            spider_result_all_df, stats_info_collect_data.spider_list
        )

        ####################################################
        # robotsレスポンスステータス解析レポートを編集
        ####################################################
        # シートを追加し選択
        self.workbook.create_sheet(title=self.ROBOTS_RESPONSE_STATUS)
        any: Any = self.workbook[self.ROBOTS_RESPONSE_STATUS]
        self.worksheet_2 = any
        self.collect_result_analysis_report_edit_header(
            self.worksheet_2, self.WORKSHEET_2_NAME, self.robots_analysis_columns_info
        )
        self.robots_warning_flg: bool = self.collect_result_analysis_report_edit_body(
            self.worksheet_2,
            stats_info_collect_data.spider_list,
            stats_info_collect_data.robots_result_df[
                StatsInfoCollectData.AGGREGATE_TYPE__SUM
            ],
            self.ROBOTS_RESPONSE_STATUS,
            self.robots_analysis_columns_info,
        )

        ####################################################
        # downloaderレスポンスステータス解析レポートを編集
        ####################################################
        # シートを追加し選択
        self.workbook.create_sheet(title=self.DOWNLOADER_RESPONSE_STATUS)
        any: Any = self.workbook[self.DOWNLOADER_RESPONSE_STATUS]
        self.worksheet_3 = any
        # downloaderレスポンスステータス統計解析レポートを編集
        self.collect_result_analysis_report_edit_header(
            self.worksheet_3,
            self.WORKSHEET_3_NAME,
            self.downloader_analysis_columns_info,
        )
        self.downloder_warning_flg: bool = (
            self.collect_result_analysis_report_edit_body(
                self.worksheet_3,
                stats_info_collect_data.spider_list,
                stats_info_collect_data.downloader_result_df[
                    StatsInfoCollectData.AGGREGATE_TYPE__SUM
                ],
                self.DOWNLOADER_RESPONSE_STATUS,
                self.downloader_analysis_columns_info,
            )
        )

    def stats_analysis_report_header(self):
        """集計結果解析レポートExcelの見出し編集"""
        self.worksheet_1.title = self.WORKSHEET_1_NAME  # シート名変更

        # 罫線定義
        side = Side(style="thin", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)
        fill1 = PatternFill(patternType="solid", fgColor="0066CC")
        fill2 = PatternFill(patternType="solid", fgColor="0099CC")

        for i, col_info in enumerate(self.stats_analysis_columns_info):
            # 見出し１行目
            any: Any = self.worksheet_1[f"{get_column_letter(i + 1)}{str(1)}"]
            head1_cell: Cell = any
            self.worksheet_1[head1_cell.coordinate] = col_info[self.HEAD1]
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(horizontal="centerContinuous")  # 選択範囲内中央寄せ

            # 見出し２行目
            any: Any = self.worksheet_1[f"{get_column_letter(i + 1)}{str(2)}"]
            head2_cell: Cell = any
            self.worksheet_1[head2_cell.coordinate] = col_info[self.HEAD2]
            head2_cell.fill = fill2
            head2_cell.border = border
            head2_cell.alignment = Alignment(horizontal="center")  # 中央寄せ

    def stats_analysis_report_body(
        self,
        spider_result_all_df: pd.DataFrame,
        spider_list: pd.Series,
    ):
        """
        統計情報解析レポート用Excelの編集。
        ワーニング有無フラグを返す。(True:有り、False:無し)
        """
        # 罫線定義
        side = Side(style="thin", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)

        # 集計期間の数を確認
        aggregate_base_term_list_count = len(
            spider_result_all_df["aggregate_base_term"].drop_duplicates()
        )

        # ワーニング有無フラグ(初期値：無し)
        warning_flg: bool = False

        # エクセルシートの編集開始行数初期化
        base_row_idx: int = 3
        # スパイダー別に解析を行う。
        for spider in spider_list:
            columns_info_by_spider = deepcopy(self.stats_analysis_columns_info)
            # スパイダー別のデータフレームを作成
            by_spider_df = spider_result_all_df.query(f'spider_name == "{spider}"')
            # データ無しの件数を確認 ※スパイダーの可動前はデータ無しとなる。
            data_none_count: int = len(
                by_spider_df.query(f'elapsed_time_seconds == ""')
            )
            nomal_count: int = aggregate_base_term_list_count - data_none_count

            # 列ごとにエクセルに編集
            for col_idx, col_info in enumerate(columns_info_by_spider):
                for row_idx, value in enumerate(by_spider_df[col_info[self.COL]]):
                    # 更新対象のセル
                    any: Any = self.worksheet_1[
                        f"{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}"
                    ]
                    target_cell: Cell = any

                    # 表示単位の切り上げ (example:b -> kb)
                    custom_value = value
                    if self.DIGIT_ADJUSTMENT in col_info:
                        if value:
                            custom_value = int(value) // col_info[self.DIGIT_ADJUSTMENT]
                            Decimal(str(custom_value)).quantize(
                                Decimal("0"), rounding=ROUND_HALF_UP
                            )

                    # 小数点以下の調整
                    if self.NUMBER_FORMAT in col_info:
                        target_cell.number_format = col_info[self.NUMBER_FORMAT]
                    else:
                        target_cell.number_format = "#,##0"

                    # 更新対象のセルに値を設定
                    self.worksheet_1[target_cell.coordinate] = custom_value

                    # 同値カラー調整
                    if self.EQUIVALENT_COLOR in col_info:
                        # 比較用の１つ上のセルと同じ値の場合は文字色を変更
                        any: Any = self.worksheet_1[
                            f"{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx - 1)}"
                        ]
                        compare_cell: Cell = any
                        if target_cell.value == compare_cell.value:
                            target_cell.font = Font(
                                color=col_info[self.EQUIVALENT_COLOR]
                            )

                    # ワーニング値チェック列の場合、ワーニング値セルの背景色を設定。
                    if self.WARNING_VALUE in col_info:
                        if value in col_info[self.WARNING_VALUE]:
                            target_cell.fill = PatternFill(
                                patternType="solid",
                                fgColor=col_info[self.WARNING_BACKGROUND_COLOR],
                            )
                            warning_flg = True

                    # ワーニング値超過チェック列の場合、ワーニング値超過セルの背景色を設定。
                    if self.WARNING_VALUE_OVER in col_info:
                        if value:
                            if int(value) > col_info[self.WARNING_VALUE_OVER]:
                                target_cell.fill = PatternFill(
                                    patternType="solid",
                                    fgColor=col_info[self.WARNING_BACKGROUND_COLOR],
                                )
                                warning_flg = True

            # １つのスパイダーに対する編集が終わった段階で、次のスパイダー用にエクセルの編集行の基準を設定する。
            base_row_idx = base_row_idx + len(by_spider_df)

        # 各明細行のセルに罫線を設定する。
        max_cell: str = get_column_letter(self.worksheet_1.max_column) + str(
            self.worksheet_1.max_row
        )  # "BC55"のようなセル番地を生成
        rows = self.worksheet_1[f"a3:{max_cell}"]
        if type(rows) is tuple:
            for cells in rows:
                cells: Any
                for cell in cells:
                    cell: Cell
                    cell.border = border

        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        # for col in ws.columns:
        for col in self.worksheet_1.iter_cols():
            max_length = 0
            column = col[0].column_letter  # 列名A,Bなどを取得
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            # 型ヒントでcolumn_dimensionsが存在しないものとみなされエラーが出るため、動的メソッドの実行形式で記述
            # getattr(ws, 'column_dimensions')()[column].width = (max_length + 2.07)
            self.worksheet_1.column_dimensions[column].width = max_length + 2.07

        # ウィンドウ枠の固定。２行２列は常に表示させる。
        self.worksheet_1.freeze_panes = "c3"

        return warning_flg

    def collect_result_analysis_report_edit_header(
        self,
        ws: Any,  # レポートの出力先のワークシート
        sheet_name: str,  # レポートの出力先のワークシートのシート名
        analysis_columns_info: list,
    ):  # レポート出力対象カラムに関する情報リスト
        """集計結果解析レポート用Excelの編集(見出し)"""

        ws.title = sheet_name  # シート名変更

        # 罫線定義
        side = Side(style="thin", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)
        # 背景色
        fill1 = PatternFill(patternType="solid", fgColor="0066CC")

        for i, col_info in enumerate(analysis_columns_info):
            # 見出し１行目
            any: Any = ws[f"{get_column_letter(i + 1)}{str(1)}"]
            head1_cell: Cell = any  # 型ヒントエラー回避用にAnyに一度入れて、改めてCellの項目に渡す
            ws[head1_cell.coordinate] = col_info[self.HEAD1]
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(horizontal="center")  # 選択範囲内中央寄せ

    def collect_result_analysis_report_edit_body(
        self,
        ws: Any,  # レポートの出力先のワークシート
        spider_list: pd.Series,  # スパイダーの一覧
        result_df: pd.DataFrame,  # mongoDBへの集計結果より作成したデータフレーム
        check_col: str,  # チェックを行いたいカラム
        analysis_columns_info: list,  # レポート出力対象カラムに関する情報リスト
    ) -> bool:
        """
        集計結果解析レポート用Excelの編集。
        ワーニング有無フラグを返す。(True:有り、False:無し)
        """

        # 罫線定義
        side = Side(style="thin", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)

        # 集計期間の数を確認
        aggregate_base_term_list_count = len(
            result_df[StatsInfoCollectData.AGGREGATE_BASE_TERM].drop_duplicates()
        )

        # ワーニング有無フラグ(初期値：無し)
        warning_flg: bool = False

        # エクセルシートの編集開始行数初期化
        base_row_idx: int = 2
        # スパイダー別に解析を行う。
        for spider in spider_list:
            columns_info_by_spider = deepcopy(analysis_columns_info)
            # スパイダー別のデータフレームを作成
            by_spider_df = result_df.query(f'spider_name == "{spider}"')
            # データ無しの件数を確認 ※スパイダーの可動前はデータ無しとなる。
            data_none_count: int = len(by_spider_df.query(f'{check_col} == ""'))
            nomal_count: int = aggregate_base_term_list_count - data_none_count

            # ステータス別の一覧を作成し、
            # ステータス別の出現件数が集計期間数と一致するか確認
            status_list_sr: pd.Series = by_spider_df[check_col].drop_duplicates()
            for status in status_list_sr:
                # データ無しは除く
                if status:
                    by_status_df = by_spider_df.query(f'{check_col} == "{status}"')
                    if nomal_count != len(by_status_df):
                        for col_info in columns_info_by_spider:
                            if col_info[self.COL] == check_col:
                                if status != 200:  # ステータスが200以外はワーニングとして記録
                                    col_info[self.WARNING_VALUE].append(status)

            # 列ごとにエクセルに編集
            for col_idx, col_info in enumerate(columns_info_by_spider):
                # データフレームの列ごとに１行ずつ処理を実施
                for row_idx, value in enumerate(by_spider_df[col_info[self.COL]]):
                    # 更新対象のセルに値を設定
                    f"{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}"
                    # any: Any = ws[get_column_letter(col_idx + 1) + str(base_row_idx + row_idx)]
                    any: Any = ws[
                        f"{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}"
                    ]
                    target_cell: Cell = any
                    ws[target_cell.coordinate] = value

                    # 同値カラー調整列の場合、更新対象セルの上と同値ならば色を変える。
                    if self.EQUIVALENT_COLOR in col_info:
                        any: Any = ws[
                            f"{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx - 1)}"
                        ]
                        compare_cell: Cell = any

                        if target_cell.value == compare_cell.value:
                            target_cell.font = Font(
                                color=col_info[self.EQUIVALENT_COLOR]
                            )
                            warning_flg = True

                    # ワーニング値チェック列の場合、ワーニング値セルの背景色を設定。
                    if self.WARNING_VALUE in col_info:
                        # if value in col_info[self.WARNING_VALUE]:
                        if (
                            value and value != "200"
                        ):  # （暫定対応）ステータスに値があり、かつ、200以外はワーニングとする。
                            target_cell.fill = PatternFill(
                                patternType="solid",
                                fgColor=col_info[self.WARNING_BACKGROUND_COLOR],
                            )
                            warning_flg = True

            # １つのスパイダーに対する編集が終わった段階で、次のスパイダー用にエクセルの編集行の基準を設定する。
            base_row_idx = base_row_idx + len(by_spider_df)

        # 各明細行のセルに罫線を設定する。
        max_cell: str = get_column_letter(ws.max_column) + str(
            ws.max_row
        )  # "BC55"のようなセル番地を生成
        rows = ws[f"a2:{max_cell}"]
        if type(rows) is tuple:
            for cells in rows:
                cells: Any
                for cell in cells:
                    cell.border = border

        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        for cols in ws.iter_cols():
            max_length = 0
            column = cols[0].column_letter  # 列名A,Bなどを取得
            for cell in cols:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            # #型ヒントでcolumn_dimensionsが存在しないものとみなされエラーが出るため、動的メソッドの実行形式で記述
            # getattr(ws, 'column_dimensions')()[column].width = (max_length + 2.07)
            ws.column_dimensions[column].width = max_length + 2.07

        # ウィンドウ枠の固定。１行２列は常に表示させる。
        ws.freeze_panes = "c2"

        return warning_flg
