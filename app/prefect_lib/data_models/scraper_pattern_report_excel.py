from typing import Any, Final

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.chart.bar_chart import BarChart
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from prefect_lib.data_models.scraper_pattern_report_data import \
    ScraperPatternReportData


class ScraperPatternReportExcel:
    """"""

    ############
    # 定数定義
    ############
    HEAD1: Final[str] = "head1"
    HEAD2: Final[str] = "head2"
    COL: Final[str] = "col"
    EQUIVALENT_COLOR: Final[str] = "equivalent_color"

    SCRAPER_PATTERN_ANALYSIS_COLUMNS_INFO: list = [
        # head1                     : 必須 : 見出し１行目
        # head2                     : 必須 : 見出し２行目
        # col                       : 必須 : データフレーム列名
        # digit_adjustment          : 任意 : 単位の調整。1000とした場合、'value/1000'となる。
        # number_format             : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # number_format             : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_value_over        : 任意 : 超過したらワーニングとする値
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        # {'head1': '集計期間', 'head2': '', 'col': 'aggregate_base_term'},
        {
            HEAD1: "スクレイパー情報",
            HEAD2: "ドメイン",
            COL: ScraperPatternReportData.DOMAIN,
            EQUIVALENT_COLOR: "bbbbbb",
        },
        {
            HEAD1: "",
            HEAD2: "アイテム",
            COL: ScraperPatternReportData.SCRAPE_ITEMS,
            EQUIVALENT_COLOR: "bbbbbb",
        },
        {HEAD1: "", HEAD2: "優先順位", COL: ScraperPatternReportData.PRIORITY},
        {HEAD1: "", HEAD2: "パターン", COL: ScraperPatternReportData.PATTERN},
        {HEAD1: "", HEAD2: "使用回数", COL: ScraperPatternReportData.COUNT_OF_USE},
    ]

    ###################
    # クラス変数定義
    ###################
    workbook: Workbook
    worksheet: Worksheet

    def __init__(self, scraper_pattern_report_data: ScraperPatternReportData) -> None:
        # スクレイパー情報集計結果報告用のワークブックの新規作成
        self.workbook = Workbook()
        any: Any = self.workbook.active  # アクティブなワークシートを選択
        self.worksheet: Worksheet = any

        # 見出し編集
        self.scraper_pattern_report_header()
        # データ編集
        self.scraper_pattern_report_body(scraper_pattern_report_data)

    def scraper_pattern_report_header(self):
        """スクレイパー情報解析レポート用Excelの見出し編集"""

        self.worksheet.title = "Scraper Info By Domain"  # シート名変更

        # 罫線定義
        side = Side(style="thin", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)
        fill1 = PatternFill(patternType="solid", fgColor="0066CC")
        fill2 = PatternFill(patternType="solid", fgColor="0099CC")

        for i, col_info in enumerate(self.SCRAPER_PATTERN_ANALYSIS_COLUMNS_INFO):
            # 見出し１行目
            any: Any = self.worksheet[f"{get_column_letter(i + 1)}{str(1)}"]
            head1_cell: Cell = any
            self.worksheet[head1_cell.coordinate] = col_info[self.HEAD1]
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(horizontal="centerContinuous")  # 選択範囲内中央寄せ

            # 見出し２行目
            any: Any = self.worksheet[f"{get_column_letter(i + 1)}{str(2)}"]
            head2_cell: Cell = any
            self.worksheet[head2_cell.coordinate] = col_info[self.HEAD2]
            head2_cell.fill = fill2
            head2_cell.border = border
            head2_cell.alignment = Alignment(horizontal="center")  # 中央寄せ

    def scraper_pattern_report_body(
        self, scraper_pattern_report_data: ScraperPatternReportData
    ):
        """
        スクレイパー情報解析レポート用Excelの編集
        """
        result_df: pd.DataFrame = scraper_pattern_report_data.result_df

        # 罫線定義
        side = Side(style="thin", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)

        # エクセルシートの編集開始行数初期化
        base_row_idx: int = 3
        # 列ごとにエクセルに編集
        for col_idx, col_info in enumerate(self.SCRAPER_PATTERN_ANALYSIS_COLUMNS_INFO):
            for row_idx, value in enumerate(result_df[col_info[self.COL]]):
                # 更新対象のセル
                any: Any = self.worksheet[
                    f"{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}"
                ]
                target_cell: Cell = any

                # 更新対象のセルに値を設定
                self.worksheet[target_cell.coordinate] = value

                # 同値カラー調整
                if self.EQUIVALENT_COLOR in col_info:
                    # 比較用の１つ上のセルと同じ値の場合は文字色を変更
                    any: Any = self.worksheet[
                        f"{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx - 1)}"
                    ]
                    compare_cell: Cell = any
                    if target_cell.value == compare_cell.value:
                        target_cell.font = Font(color=col_info[self.EQUIVALENT_COLOR])

        # 各明細行のセルに罫線を設定する。
        max_cell: str = get_column_letter(self.worksheet.max_column) + str(
            self.worksheet.max_row
        )  # "BC55"のようなセル番地を生成
        rows = self.worksheet[f"a3:{max_cell}"]
        if type(rows) is tuple:
            for cells in rows:
                cells: Any
                for cell in cells:
                    cell.border = border

        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        for col in self.worksheet.iter_cols():
            max_length = 0
            column = col[0].column_letter  # 列名A,Bなどを取得
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            # 型ヒントでcolumn_dimensionsが存在しないものとみなされエラーが出るため、動的メソッドの実行形式で記述
            # getattr(worksheet, 'column_dimensions')()[column].width = (max_length + 2.2)
            self.worksheet.column_dimensions[column].width = max_length + 5.0

        # ウィンドウ枠の固定。２行２列は常に表示させる。
        self.worksheet.freeze_panes = "c3"
